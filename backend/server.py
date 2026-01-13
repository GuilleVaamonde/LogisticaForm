from fastapi import FastAPI, APIRouter, HTTPException, Depends, status, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from fastapi.security import HTTPBearer, HTTPAuthorizationCredentials
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
from pathlib import Path
from pydantic import BaseModel, Field, ConfigDict
from typing import List, Optional
import uuid
from datetime import datetime, timezone, timedelta
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
import jwt
import bcrypt
import base64


ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env')

# MongoDB connection
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# JWT Config
JWT_SECRET = os.environ['JWT_SECRET']
JWT_ALGORITHM = "HS256"
JWT_EXPIRATION_HOURS = 24

# Frontend URL for tracking links
FRONTEND_URL = os.environ.get('FRONTEND_URL', 'https://cliente-form.preview.emergentagent.com')

# Create the main app
app = FastAPI()

# Create routers
api_router = APIRouter(prefix="/api")
auth_router = APIRouter(prefix="/api/auth", tags=["auth"])
users_router = APIRouter(prefix="/api/users", tags=["users"])
envios_router = APIRouter(prefix="/api/envios", tags=["envios"])
messages_router = APIRouter(prefix="/api/messages", tags=["messages"])
tracking_router = APIRouter(prefix="/api/tracking", tags=["tracking"])

security = HTTPBearer()

# Uruguay departments
DEPARTAMENTOS_URUGUAY = [
    "Artigas", "Canelones", "Cerro Largo", "Colonia", "Durazno",
    "Flores", "Florida", "Lavalleja", "Maldonado", "Montevideo",
    "Paysandú", "Río Negro", "Rivera", "Rocha", "Salto",
    "San José", "Soriano", "Tacuarembó", "Treinta y Tres"
]

# Motivos de envío
MOTIVOS_ENVIO = ["Entrega", "Retiro y Entrega", "Retiro"]

# Estados de envío
ESTADOS_ENVIO = ["Ingresada", "Asignado a courier", "Entregado"]

# Roles
ROLES = ["admin", "agente", "repartidor"]


# ============== MODELS ==============

class UserBase(BaseModel):
    username: str = Field(..., min_length=3)
    nombre: str = Field(..., min_length=2)
    rol: str = Field(..., description="admin, agente, repartidor")


class UserCreate(UserBase):
    password: str = Field(..., min_length=4)


class UserResponse(UserBase):
    model_config = ConfigDict(extra="ignore")
    id: str
    activo: bool = True
    created_at: str


class UserLogin(BaseModel):
    username: str
    password: str


class TokenResponse(BaseModel):
    access_token: str
    token_type: str = "bearer"
    user: UserResponse


class EnvioBase(BaseModel):
    ticket: str = Field(..., min_length=1, description="Número de ticket para rastreo")
    calle: str = Field(..., min_length=1, description="Nombre de la calle")
    numero: str = Field(..., min_length=1, description="Número de casa")
    apto: Optional[str] = Field(default="", description="Número de apartamento")
    esquina: Optional[str] = Field(default="", description="Esquina de referencia")
    motivo: str = Field(..., description="Motivo del envío")
    departamento: str = Field(..., description="Departamento de Uruguay")
    comentarios: Optional[str] = Field(default="", description="Comentarios adicionales")
    telefono: str = Field(..., min_length=1, description="Número de teléfono")
    contacto: str = Field(..., min_length=1, description="Nombre de contacto")


class EnvioCreate(EnvioBase):
    pass


class EnvioUpdate(BaseModel):
    ticket: Optional[str] = None
    calle: Optional[str] = None
    numero: Optional[str] = None
    apto: Optional[str] = None
    esquina: Optional[str] = None
    motivo: Optional[str] = None
    departamento: Optional[str] = None
    comentarios: Optional[str] = None
    telefono: Optional[str] = None
    contacto: Optional[str] = None


class EstadoHistorial(BaseModel):
    estado: str
    fecha: str
    usuario_id: str
    usuario_nombre: str
    receptor_nombre: Optional[str] = None
    receptor_cedula: Optional[str] = None
    imagen_url: Optional[str] = None


class EnvioResponse(BaseModel):
    id: str
    ticket: str
    calle: str
    numero: str
    apto: str
    esquina: str
    motivo: str
    departamento: str
    comentarios: str
    telefono: str
    contacto: str
    fecha_carga: str
    estado: str
    historial_estados: List[EstadoHistorial]
    creado_por: str
    creado_por_nombre: str


class CambioEstadoRequest(BaseModel):
    nuevo_estado: str
    receptor_nombre: Optional[str] = None
    receptor_cedula: Optional[str] = None


class MessageLog(BaseModel):
    id: str
    envio_id: str
    ticket: str
    telefono: str
    mensaje: str
    estado: str
    fecha: str
    enviado: bool = False  # False = simulado, True = enviado real


class EnvioFilters(BaseModel):
    departamento: Optional[str] = None
    motivo: Optional[str] = None
    fecha_desde: Optional[str] = None
    fecha_hasta: Optional[str] = None
    estado: Optional[str] = None


# ============== HELPERS ==============

def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode(), bcrypt.gensalt()).decode()


def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode(), hashed.encode())


def create_token(user_id: str, username: str, rol: str) -> str:
    payload = {
        "sub": user_id,
        "username": username,
        "rol": rol,
        "exp": datetime.now(timezone.utc) + timedelta(hours=JWT_EXPIRATION_HOURS)
    }
    return jwt.encode(payload, JWT_SECRET, algorithm=JWT_ALGORITHM)


async def get_current_user(credentials: HTTPAuthorizationCredentials = Depends(security)):
    try:
        token = credentials.credentials
        payload = jwt.decode(token, JWT_SECRET, algorithms=[JWT_ALGORITHM])
        user_id = payload.get("sub")
        
        user = await db.users.find_one({"id": user_id, "activo": True}, {"_id": 0})
        if not user:
            raise HTTPException(status_code=401, detail="Usuario no encontrado")
        
        return user
    except jwt.ExpiredSignatureError:
        raise HTTPException(status_code=401, detail="Token expirado")
    except jwt.InvalidTokenError:
        raise HTTPException(status_code=401, detail="Token inválido")


def require_role(*roles):
    async def role_checker(user: dict = Depends(get_current_user)):
        if user["rol"] not in roles:
            raise HTTPException(
                status_code=403, 
                detail=f"Acceso denegado. Se requiere rol: {', '.join(roles)}"
            )
        return user
    return role_checker


async def log_whatsapp_message(envio_id: str, ticket: str, telefono: str, mensaje: str, estado: str):
    """Log WhatsApp message (simulated for now, ready for WhatsApp Business API)"""
    message_log = {
        "id": str(uuid.uuid4()),
        "envio_id": envio_id,
        "ticket": ticket,
        "telefono": telefono,
        "mensaje": mensaje,
        "estado": estado,
        "fecha": datetime.now(timezone.utc).isoformat(),
        "enviado": False  # Simulated
    }
    await db.message_logs.insert_one(message_log)
    return message_log


def create_excel_workbook(envios: List[dict]) -> BytesIO:
    """Create an Excel workbook with envio data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Envíos"
    
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="C91A25", end_color="C91A25", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = Border(
        left=Side(style='thin', color='E2E8F0'),
        right=Side(style='thin', color='E2E8F0'),
        top=Side(style='thin', color='E2E8F0'),
        bottom=Side(style='thin', color='E2E8F0')
    )
    
    headers = [
        "Ticket", "Estado", "Fecha/Hora", "Contacto", "Teléfono", "Calle", 
        "Número", "Apto", "Esquina", "Departamento", "Motivo", "Comentarios"
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    for row, envio in enumerate(envios, 2):
        data = [
            envio.get('ticket', ''),
            envio.get('estado', 'Ingresada'),
            envio.get('fecha_carga', ''),
            envio.get('contacto', ''),
            envio.get('telefono', ''),
            envio.get('calle', ''),
            envio.get('numero', ''),
            envio.get('apto', ''),
            envio.get('esquina', ''),
            envio.get('departamento', ''),
            envio.get('motivo', ''),
            envio.get('comentarios', '')
        ]
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    column_widths = [15, 18, 22, 20, 15, 25, 10, 10, 20, 15, 18, 30]
    for i, width in enumerate(column_widths, 1):
        ws.column_dimensions[chr(64 + i)].width = width
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


# ============== AUTH ROUTES ==============

@auth_router.post("/login", response_model=TokenResponse)
async def login(credentials: UserLogin):
    user = await db.users.find_one({"username": credentials.username, "activo": True}, {"_id": 0})
    
    if not user or not verify_password(credentials.password, user["password"]):
        raise HTTPException(status_code=401, detail="Credenciales inválidas")
    
    token = create_token(user["id"], user["username"], user["rol"])
    
    user_response = UserResponse(
        id=user["id"],
        username=user["username"],
        nombre=user["nombre"],
        rol=user["rol"],
        activo=user["activo"],
        created_at=user["created_at"]
    )
    
    return TokenResponse(access_token=token, user=user_response)


@auth_router.get("/me", response_model=UserResponse)
async def get_me(user: dict = Depends(get_current_user)):
    return UserResponse(**user)


# ============== USER ROUTES ==============

@users_router.post("", response_model=UserResponse)
async def create_user(
    user_data: UserCreate,
    current_user: dict = Depends(require_role("admin"))
):
    if user_data.rol not in ROLES:
        raise HTTPException(status_code=400, detail=f"Rol inválido. Opciones: {', '.join(ROLES)}")
    
    existing = await db.users.find_one({"username": user_data.username})
    if existing:
        raise HTTPException(status_code=400, detail="El usuario ya existe")
    
    user = {
        "id": str(uuid.uuid4()),
        "username": user_data.username,
        "password": hash_password(user_data.password),
        "nombre": user_data.nombre,
        "rol": user_data.rol,
        "activo": True,
        "created_at": datetime.now(timezone.utc).isoformat()
    }
    
    await db.users.insert_one(user)
    
    return UserResponse(
        id=user["id"],
        username=user["username"],
        nombre=user["nombre"],
        rol=user["rol"],
        activo=user["activo"],
        created_at=user["created_at"]
    )


@users_router.get("", response_model=List[UserResponse])
async def get_users(current_user: dict = Depends(require_role("admin"))):
    users = await db.users.find({"activo": True}, {"_id": 0, "password": 0}).to_list(1000)
    return [UserResponse(**u) for u in users]


@users_router.delete("/{user_id}")
async def delete_user(user_id: str, current_user: dict = Depends(require_role("admin"))):
    if user_id == current_user["id"]:
        raise HTTPException(status_code=400, detail="No puedes eliminarte a ti mismo")
    
    result = await db.users.update_one(
        {"id": user_id},
        {"$set": {"activo": False}}
    )
    
    if result.modified_count == 0:
        raise HTTPException(status_code=404, detail="Usuario no encontrado")
    
    return {"message": "Usuario eliminado"}


# ============== ENVIOS ROUTES ==============

@api_router.get("/")
async def root():
    return {"message": "API de Gestión de Envíos - Uruguay"}


@api_router.get("/departamentos")
async def get_departamentos():
    return {"departamentos": DEPARTAMENTOS_URUGUAY}


@api_router.get("/motivos")
async def get_motivos():
    return {"motivos": MOTIVOS_ENVIO}


@api_router.get("/estados")
async def get_estados():
    return {"estados": ESTADOS_ENVIO}


@envios_router.post("", response_model=EnvioResponse)
async def create_envio(
    envio_data: EnvioCreate,
    current_user: dict = Depends(require_role("admin", "agente"))
):
    if envio_data.departamento not in DEPARTAMENTOS_URUGUAY:
        raise HTTPException(status_code=400, detail="Departamento inválido")
    
    if envio_data.motivo not in MOTIVOS_ENVIO:
        raise HTTPException(status_code=400, detail="Motivo inválido")
    
    now = datetime.now(timezone.utc).isoformat()
    
    historial_inicial = {
        "estado": "Ingresada",
        "fecha": now,
        "usuario_id": current_user["id"],
        "usuario_nombre": current_user["nombre"]
    }
    
    envio = {
        "id": str(uuid.uuid4()),
        **envio_data.model_dump(),
        "fecha_carga": now,
        "estado": "Ingresada",
        "historial_estados": [historial_inicial],
        "creado_por": current_user["id"],
        "creado_por_nombre": current_user["nombre"]
    }
    
    await db.envios.insert_one(envio)
    
    return EnvioResponse(**envio)


@envios_router.get("", response_model=List[EnvioResponse])
async def get_envios(
    limit: int = 100,
    skip: int = 0,
    departamento: Optional[str] = None,
    motivo: Optional[str] = None,
    estado: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    
    if departamento:
        query["departamento"] = departamento
    if motivo:
        query["motivo"] = motivo
    if estado:
        query["estado"] = estado
    if fecha_desde:
        query["fecha_carga"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha_carga" in query:
            query["fecha_carga"]["$lte"] = fecha_hasta
        else:
            query["fecha_carga"] = {"$lte": fecha_hasta}
    
    envios = await db.envios.find(
        query, 
        {"_id": 0}
    ).sort("fecha_carga", -1).skip(skip).limit(limit).to_list(limit)
    
    return [EnvioResponse(**e) for e in envios]


@envios_router.get("/count")
async def get_envios_count(
    departamento: Optional[str] = None,
    motivo: Optional[str] = None,
    estado: Optional[str] = None,
    current_user: dict = Depends(get_current_user)
):
    query = {}
    if departamento:
        query["departamento"] = departamento
    if motivo:
        query["motivo"] = motivo
    if estado:
        query["estado"] = estado
    
    count = await db.envios.count_documents(query)
    return {"count": count}


@envios_router.get("/export/excel")
async def export_all_envios_excel(
    departamento: Optional[str] = None,
    motivo: Optional[str] = None,
    estado: Optional[str] = None,
    fecha_desde: Optional[str] = None,
    fecha_hasta: Optional[str] = None,
    current_user: dict = Depends(require_role("admin", "agente"))
):
    query = {}
    if departamento:
        query["departamento"] = departamento
    if motivo:
        query["motivo"] = motivo
    if estado:
        query["estado"] = estado
    if fecha_desde:
        query["fecha_carga"] = {"$gte": fecha_desde}
    if fecha_hasta:
        if "fecha_carga" in query:
            query["fecha_carga"]["$lte"] = fecha_hasta
        else:
            query["fecha_carga"] = {"$lte": fecha_hasta}
    
    envios = await db.envios.find(query, {"_id": 0}).sort("fecha_carga", -1).to_list(1000)
    
    if not envios:
        raise HTTPException(status_code=404, detail="No hay envíos para exportar")
    
    excel_file = create_excel_workbook(envios)
    filename = f"envios_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@envios_router.get("/{envio_id}", response_model=EnvioResponse)
async def get_envio(envio_id: str, current_user: dict = Depends(get_current_user)):
    envio = await db.envios.find_one({"id": envio_id}, {"_id": 0})
    if not envio:
        raise HTTPException(status_code=404, detail="Envío no encontrado")
    return EnvioResponse(**envio)


@envios_router.put("/{envio_id}", response_model=EnvioResponse)
async def update_envio(
    envio_id: str,
    envio_data: EnvioUpdate,
    current_user: dict = Depends(require_role("admin", "agente"))
):
    envio = await db.envios.find_one({"id": envio_id}, {"_id": 0})
    if not envio:
        raise HTTPException(status_code=404, detail="Envío no encontrado")
    
    update_data = {k: v for k, v in envio_data.model_dump().items() if v is not None}
    
    if update_data:
        await db.envios.update_one({"id": envio_id}, {"$set": update_data})
    
    updated_envio = await db.envios.find_one({"id": envio_id}, {"_id": 0})
    return EnvioResponse(**updated_envio)


@envios_router.patch("/{envio_id}/estado", response_model=EnvioResponse)
async def cambiar_estado(
    envio_id: str,
    cambio: CambioEstadoRequest,
    current_user: dict = Depends(get_current_user)
):
    if cambio.nuevo_estado not in ESTADOS_ENVIO:
        raise HTTPException(status_code=400, detail=f"Estado inválido. Opciones: {', '.join(ESTADOS_ENVIO)}")
    
    envio = await db.envios.find_one({"id": envio_id}, {"_id": 0})
    if not envio:
        raise HTTPException(status_code=404, detail="Envío no encontrado")
    
    # Validate state transition
    current_state = envio["estado"]
    valid_transitions = {
        "Ingresada": ["Asignado a courier"],
        "Asignado a courier": ["Entregado"],
        "Entregado": []
    }
    
    if cambio.nuevo_estado not in valid_transitions.get(current_state, []):
        raise HTTPException(
            status_code=400, 
            detail=f"No se puede cambiar de '{current_state}' a '{cambio.nuevo_estado}'"
        )
    
    # For "Entregado" state, require receptor info
    if cambio.nuevo_estado == "Entregado":
        if not cambio.receptor_nombre or not cambio.receptor_cedula:
            raise HTTPException(
                status_code=400, 
                detail="Se requiere nombre y cédula del receptor para marcar como entregado"
            )
    
    now = datetime.now(timezone.utc).isoformat()
    
    nuevo_historial = {
        "estado": cambio.nuevo_estado,
        "fecha": now,
        "usuario_id": current_user["id"],
        "usuario_nombre": current_user["nombre"],
        "receptor_nombre": cambio.receptor_nombre,
        "receptor_cedula": cambio.receptor_cedula
    }
    
    await db.envios.update_one(
        {"id": envio_id},
        {
            "$set": {"estado": cambio.nuevo_estado},
            "$push": {"historial_estados": nuevo_historial}
        }
    )
    
    # Send WhatsApp message (simulated)
    mensaje = None
    if cambio.nuevo_estado == "Asignado a courier":
        mensaje = f"🚚 Tu pedido con ticket {envio['ticket']} fue asignado a un cadete. Pronto llegará a tu dirección."
    elif cambio.nuevo_estado == "Entregado":
        mensaje = f"✅ Tu pedido con ticket {envio['ticket']} ha sido entregado. Recibido por: {cambio.receptor_nombre}. ¡Gracias por confiar en nosotros!"
    
    if mensaje:
        await log_whatsapp_message(
            envio_id=envio_id,
            ticket=envio["ticket"],
            telefono=envio["telefono"],
            mensaje=mensaje,
            estado=cambio.nuevo_estado
        )
    
    updated_envio = await db.envios.find_one({"id": envio_id}, {"_id": 0})
    return EnvioResponse(**updated_envio)


@envios_router.get("/{envio_id}/excel")
async def export_single_envio_excel(
    envio_id: str,
    current_user: dict = Depends(require_role("admin", "agente"))
):
    envio = await db.envios.find_one({"id": envio_id}, {"_id": 0})
    
    if not envio:
        raise HTTPException(status_code=404, detail="Envío no encontrado")
    
    excel_file = create_excel_workbook([envio])
    filename = f"envio_{envio.get('ticket', envio_id)}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )


@envios_router.delete("/{envio_id}")
async def delete_envio(
    envio_id: str,
    current_user: dict = Depends(require_role("admin", "agente"))
):
    result = await db.envios.delete_one({"id": envio_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="Envío no encontrado")
    return {"message": "Envío eliminado exitosamente"}


# ============== MESSAGES ROUTES ==============

@messages_router.get("", response_model=List[MessageLog])
async def get_message_logs(
    limit: int = 50,
    current_user: dict = Depends(require_role("admin"))
):
    messages = await db.message_logs.find(
        {}, {"_id": 0}
    ).sort("fecha", -1).limit(limit).to_list(limit)
    return messages


@messages_router.get("/{envio_id}", response_model=List[MessageLog])
async def get_envio_messages(
    envio_id: str,
    current_user: dict = Depends(get_current_user)
):
    messages = await db.message_logs.find(
        {"envio_id": envio_id}, {"_id": 0}
    ).sort("fecha", -1).to_list(100)
    return messages


# ============== INIT ADMIN ==============

@app.on_event("startup")
async def create_default_admin():
    admin = await db.users.find_one({"username": "admin"})
    if not admin:
        admin_user = {
            "id": str(uuid.uuid4()),
            "username": "admin",
            "password": hash_password("admin123"),
            "nombre": "Administrador",
            "rol": "admin",
            "activo": True,
            "created_at": datetime.now(timezone.utc).isoformat()
        }
        await db.users.insert_one(admin_user)
        logging.info("Admin user created: admin / admin123")


# Include routers
app.include_router(api_router)
app.include_router(auth_router)
app.include_router(users_router)
app.include_router(envios_router)
app.include_router(messages_router)

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"],
    allow_headers=["*"],
)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)s - %(name)s - %(levelname)s - %(message)s'
)
logger = logging.getLogger(__name__)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
